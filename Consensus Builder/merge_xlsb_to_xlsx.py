# -*- coding: utf-8 -*-
"""
ECFC Consensus xlsb -> xlsx 병합 스크립트
============================================
- 회사 PC에서 xlsb가 안 열리는 환경 우회용.
- xlsb에는 최근 약 60일치만 들어있고, xlsx에는 누적 시계열이 들어있음.
- 두 파일의 양식(시트명/헤더 R1~R13/티커 컬럼)이 동일하다는 전제 하에,
  xlsx의 마지막 유효 날짜 다음날부터의 xlsb 행만 이어붙여서
  새 파일 `ECFC_..._YYYYMMDD.xlsx`로 저장.

쉽게 비유:
- xlsx = 두꺼운 가계부 본책 (작년부터 쭉)
- xlsb = 최근 60일치 영수증 묶음
- 본책 맨 뒤 페이지(=마지막 기록일)부터 그 다음 날 영수증만 풀로 붙여 넣고,
  본책 이름을 오늘 날짜로 갈아서 새 카피본을 만든다.

사용법:
    python "merge_xlsb_to_xlsx.py"            # 오늘 날짜 사용
    python "merge_xlsb_to_xlsx.py" 20260519   # 명시적 날짜
"""

import glob
import os
import sys
from datetime import datetime, timedelta

from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb

# ============================================================
# CONFIG
# ============================================================
# 스크립트가 있는 폴더를 자동 기준으로 사용 (Windows/Linux 양쪽 모두 OK).
# 다른 폴더에서 돌리고 싶으면 환경변수 CONSENSUS_DIR로 override 가능.
BASE_DIR = os.environ.get(
    'CONSENSUS_DIR',
    os.path.dirname(os.path.abspath(__file__)),
)
HISTORY_DIR = os.path.join(BASE_DIR, 'history')

# (원본 xlsx, xlsb, 출력 파일 prefix). 모두 BASE_DIR 안에 있다고 가정.
TARGETS = [
    {
        'xlsx': 'ECFC_Growth Consesus_수정.xlsx',
        'xlsb': 'ECFC_Growth Consesus_수정.xlsb',
        'out_prefix': 'ECFC_Growth Consesus',
    },
    {
        'xlsx': 'ECFC_Inflation Consesus_수정.xlsx',
        'xlsb': 'ECFC_Inflation Consesus_수정.xlsb',
        'out_prefix': 'ECFC_Inflation Consesus',
    },
]

HEADER_ROWS = 13  # R14부터 실제 일자별 데이터가 시작
# ============================================================


def serial_to_datetime(value):
    """엑셀 시리얼 number -> datetime. datetime이면 그대로 반환."""
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime(1899, 12, 30) + timedelta(days=float(value))
    except (TypeError, ValueError):
        return None


def read_xlsb_sheet(xlsb_path, sheet_name):
    """xlsb의 한 시트를 2D list로 반환."""
    with open_xlsb(xlsb_path) as wb:
        with wb.get_sheet(sheet_name) as sh:
            return [[c.v for c in row] for row in sh.rows()]


def find_xlsx_last_data_row(ws, header_rows=HEADER_ROWS):
    """A열 기준으로 마지막 유효 데이터 행(1-indexed)을 반환."""
    last_row = header_rows
    for r in range(header_rows + 1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v in (None, ''):
            # 다음 5행 연속 비면 종료 (단일 빈 행 방어)
            empty_run = True
            for rr in range(r, min(r + 5, ws.max_row + 1)):
                if ws.cell(rr, 1).value not in (None, ''):
                    empty_run = False
                    break
            if empty_run:
                break
            else:
                continue
        last_row = r
    return last_row


def find_last_data_col(ws, header_rows=HEADER_ROWS):
    """R12(티커 행 = header_rows-1) 기준으로 마지막 비어있지 않은 컬럼 인덱스."""
    ticker_row = header_rows - 1
    last_col = 1
    for c in range(2, ws.max_column + 1):
        if ws.cell(ticker_row, c).value not in (None, ''):
            last_col = c
    return max(last_col, 2)


def coverage_of(path):
    """xlsx 첫 시트의 (마지막 데이터 날짜, 데이터 행수). 못 읽으면 (None, 0)."""
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        dates = [serial_to_datetime(r[0])
                 for r in ws.iter_rows(min_row=HEADER_ROWS + 1, max_col=1, values_only=True)]
        wb.close()
        dates = [d for d in dates if d]
        return (dates[-1], len(dates)) if dates else (None, 0)
    except Exception:
        return (None, 0)


def pick_base(target, base_dir=BASE_DIR, out_dir=HISTORY_DIR):
    """베이스로 쓸 xlsx 를 고른다.

    원본 '_수정.xlsx' 는 갱신되지 않아 시간이 지나면 뒤처진다. xlsb 에는 최근
    60일치만 들어있어서, 베이스가 그 창보다 오래되면 사이 기간이 통째로 비었다.
    (실제로 베이스가 2026-04-07 에 멈춰 있고 xlsb 가 06-26 부터라 80일이 빠졌다)
    그래서 직전 산출물(history) 중 가장 멀리 온 것을 베이스로 삼아 이어붙인다.
    같은 날짜까지 왔으면 행이 많은 쪽 — 중간이 빈 파일을 피한다.
    """
    cands = [os.path.join(base_dir, target['xlsx'])]
    cands += glob.glob(os.path.join(out_dir, target['out_prefix'] + '_*.xlsx'))
    best, best_key = None, None
    for p in cands:
        if not os.path.exists(p):
            continue
        last, rows = coverage_of(p)
        if last is None:
            continue
        key = (last, rows)
        if best_key is None or key > best_key:
            best, best_key = p, key
    if best is None:
        raise FileNotFoundError(target['xlsx'])
    return best


def warn_if_gap(ws, max_gap_days=7):
    """병합 결과에 비정상적으로 긴 공백이 있으면 알린다 (조용히 넘어가지 않도록)."""
    dates = [serial_to_datetime(r[0])
             for r in ws.iter_rows(min_row=HEADER_ROWS + 1, max_col=1, values_only=True)]
    dates = [d for d in dates if d]
    gaps = [(a, b) for a, b in zip(dates, dates[1:]) if (b - a).days > max_gap_days]
    for a, b in gaps:
        print("  [!] 공백 " + str((b - a).days) + "일: "
              + str(a.date()) + " -> " + str(b.date()))
    return gaps


def merge_one(target, out_date_str, base_dir=BASE_DIR, out_dir=HISTORY_DIR,
              base_override=None):
    xlsb_path = os.path.join(base_dir, target['xlsb'])
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, target['out_prefix'] + '_' + out_date_str + '.xlsx')

    xlsx_path = base_override or pick_base(target, base_dir, out_dir)
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(xlsx_path)
    if not os.path.exists(xlsb_path):
        raise FileNotFoundError(xlsb_path)

    _last, _rows = coverage_of(xlsx_path)
    print("[Load] " + os.path.basename(xlsx_path)
          + " (베이스: ~" + (str(_last.date()) if _last else "?")
          + ", " + str(_rows) + "행)")
    wb = load_workbook(xlsx_path)

    total_appended = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        last_row = find_xlsx_last_data_row(ws)
        last_col = find_last_data_col(ws)

        last_date_raw = ws.cell(last_row, 1).value
        last_date = last_date_raw if isinstance(last_date_raw, datetime) else serial_to_datetime(last_date_raw)
        if last_date is None:
            print("  [" + sn + "] WARN: xlsx 마지막 날짜를 인식 못함. 스킵.")
            continue

        template_cells = [ws.cell(last_row, c) for c in range(1, last_col + 1)]

        try:
            xlsb_rows = read_xlsb_sheet(xlsb_path, sn)
        except Exception as e:
            print("  [" + sn + "] WARN: xlsb 시트 읽기 실패 (" + str(e) + "). 스킵.")
            continue

        data_block = xlsb_rows[HEADER_ROWS:]

        appended = 0
        new_row = last_row + 1
        for row_data in data_block:
            if not row_data:
                continue
            d_raw = row_data[0]
            if d_raw is None or d_raw == '':
                break
            d = d_raw if isinstance(d_raw, datetime) else serial_to_datetime(d_raw)
            if d is None:
                continue
            if d <= last_date:
                continue

            ws.cell(new_row, 1).value = d
            try:
                ws.cell(new_row, 1).number_format = template_cells[0].number_format
            except Exception:
                pass

            for c in range(2, last_col + 1):
                v = row_data[c - 1] if (c - 1) < len(row_data) else None
                ws.cell(new_row, c).value = v
                try:
                    ws.cell(new_row, c).number_format = template_cells[c - 1].number_format
                except Exception:
                    pass

            new_row += 1
            appended += 1

        total_appended += appended
        new_last = ws.cell(new_row - 1, 1).value if appended else last_date
        new_last_disp = new_last.date() if isinstance(new_last, datetime) else new_last
        print("  [" + sn + "] last_row=" + str(last_row) +
              " (" + str(last_date.date()) + ") -> +" + str(appended) +
              " rows, new last = " + str(new_last_disp))
        warn_if_gap(ws)

    print("[Save] " + os.path.basename(out_path) + " (총 추가 행: " + str(total_appended) + ")")
    wb.save(out_path)
    return out_path


def main():
    argv = sys.argv[1:]

    # --base <경로|YYYYMMDD> : 베이스를 직접 지정 (자동 선택이 잘못 골랐을 때 복구용).
    # 산출물 하나가 망가져서 그게 최신으로 뽑히는 상황을 사람이 되돌릴 수 있게 둔다.
    base_arg = None
    if '--base' in argv:
        i = argv.index('--base')
        base_arg = argv[i + 1] if i + 1 < len(argv) else None
        del argv[i:i + 2]
        if not base_arg:
            raise SystemExit("--base 뒤에 경로 또는 YYYYMMDD 를 주세요.")

    if argv and argv[0].isdigit() and len(argv[0]) == 8:
        date_str = argv[0]
    else:
        date_str = datetime.now().strftime('%Y%m%d')
    print("=== Output date suffix: " + date_str + " ===")

    outs = []
    for t in TARGETS:
        try:
            override = None
            if base_arg:
                override = (os.path.join(HISTORY_DIR,
                                         t['out_prefix'] + '_' + base_arg + '.xlsx')
                            if base_arg.isdigit() and len(base_arg) == 8 else base_arg)
            outs.append(merge_one(t, date_str, base_override=override))
        except Exception as e:
            print("!! 실패: " + t['xlsx'] + " -> " + str(e))

    print("\n=== Done ===")
    for o in outs:
        print("  - " + o)


if __name__ == '__main__':
    main()
