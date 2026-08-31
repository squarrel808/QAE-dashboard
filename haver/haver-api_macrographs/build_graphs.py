# build_graphs.py
"""
macro_raw_data.xlsx(레벨 원계열) + charts_config.xlsx(차트 구성)
→ Macro_Graphs_Haver.xlsx 생성.

출력 구조 (국가별 시트 2개):
  - {CC}        : 차트 시트. 원본 대시보드의 차트 묶음을 2열 격자로 배치
  - {CC}_data   : 변환 적용된 데이터 (1행 라벨, 2행 Haver 티커+변환, 3행부터 값)

변환(transform) 규칙 — charts_config.xlsx Series 시트의 transform 열:
  level        : 원계열 그대로
  yoy          : 전년동기대비 % (주기별 periods: M=12, Q=4, W=52, D=261)
  mom          : 전기대비 %
  diff         : 전기 차분
  3mma         : 3기 이동평균
  'yoy+3mma' 처럼 +로 이어 쓰면 순서대로 적용
"""
import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
CONFIG_FILE = BASE / "charts_config.xlsx"
RAW_FILE = BASE / "macro_raw_data.xlsx"
OUTPUT_FILE = BASE / "Macro_Graphs_Haver.xlsx"

CHART_W_COLS = 8    # 차트 하나가 차지하는 폭(열)
CHART_H_ROWS = 15   # 차트 하나가 차지하는 높이(행)
COUNTRY_ORDER = ["UK", "DE", "FR", "IT", "CA", "JP", "AU"]

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("macrographs")

YOY_PERIODS = {"M": 12, "Q": 4, "W": 52, "D": 261, "A": 1, "Y": 1}


def ticker_to_pk(ticker):
    s = str(ticker).strip()
    if "@" in s:
        code, db = s.split("@", 1)
        return f"{db.strip().lower()}:{code.strip().lower()}"
    return s.lower()


def apply_transform(s: pd.Series, transform: str, freq: str) -> pd.Series:
    out = s.dropna()
    for step in str(transform or "level").split("+"):
        step = step.strip().lower()
        if step in ("", "level"):
            continue
        elif step == "yoy":
            out = out.pct_change(YOY_PERIODS.get(str(freq).upper()[:1], 12)) * 100.0
        elif step == "mom":
            out = out.pct_change(1) * 100.0
        elif step == "diff":
            out = out.diff(1)
        elif step == "3mma":
            out = out.rolling(3, min_periods=3).mean()
        else:
            log.warning("알 수 없는 transform '%s' — level로 처리", step)
    return out


def main():
    charts_df = pd.read_excel(CONFIG_FILE, sheet_name="Charts")
    series_df = pd.read_excel(CONFIG_FILE, sheet_name="Series")
    wide = pd.read_excel(RAW_FILE, sheet_name="Wide", index_col=0, parse_dates=True)
    meta = pd.read_excel(RAW_FILE, sheet_name="Metadata")
    freq_map = dict(zip(meta.get("ticker_pk", pd.Series(dtype=str)).astype(str),
                        meta.get("frequency", pd.Series(dtype=str)).astype(str)))

    wb = Workbook()
    wb.remove(wb.active)
    n_charts_total = 0
    skipped_series = []

    countries = [c for c in COUNTRY_ORDER if c in set(charts_df["country"])]
    for cc in countries:
        cc_charts = charts_df[charts_df["country"] == cc].sort_values("chart_id")
        cc_series = series_df[series_df["country"] == cc]

        # ── 1) 국가 패널: (haver, transform) 유니크 시리즈 → 데이터 시트 컬럼 ──
        panel_cols = []      # [(colkey, label, haver, transform, series)]
        colkey_index = {}
        for _, row in cc_series.iterrows():
            hv = str(row["haver_ticker"]).strip()
            tr = str(row.get("transform", "level")).strip() or "level"
            key = (hv, tr)
            if key in colkey_index or not hv or hv.lower() == "nan":
                continue
            pk = ticker_to_pk(hv)
            if pk not in wide.columns:
                skipped_series.append((cc, hv, "raw 데이터 없음"))
                continue
            vals = apply_transform(wide[pk], tr, freq_map.get(pk, "M"))
            colkey_index[key] = len(panel_cols)
            panel_cols.append((key, str(row.get("label", hv)), hv, tr, vals))

        if not panel_cols:
            continue

        panel = pd.DataFrame({i: s for i, (_, _, _, _, s) in enumerate(panel_cols)})
        panel = panel.sort_index()
        panel = panel.dropna(how="all")

        # ── 2) 데이터 시트 쓰기 ──
        ws_d = wb.create_sheet(f"{cc}_data")
        ws_d["A1"] = "label"
        ws_d["A2"] = "haver (transform)"
        for j, (key, label, hv, tr, _) in enumerate(panel_cols):
            col = j + 2
            ws_d.cell(row=1, column=col, value=label)
            ws_d.cell(row=2, column=col, value=f"{hv} ({tr})")
        for i, (dt, row_vals) in enumerate(panel.iterrows()):
            r = i + 3
            ws_d.cell(row=r, column=1, value=dt.to_pydatetime()).number_format = "yyyy-mm"
            for j in range(len(panel_cols)):
                v = row_vals[j]
                if pd.notna(v):
                    ws_d.cell(row=r, column=j + 2, value=float(v))
        ws_d.freeze_panes = "B3"
        ws_d.column_dimensions["A"].width = 11
        n_rows = len(panel)
        first_data_row, last_data_row = 3, 2 + n_rows

        # ── 3) 차트 시트 ──
        ws_c = wb.create_sheet(cc)
        pos = 0
        for _, chrow in cc_charts.iterrows():
            chart_id = chrow["chart_id"]
            rows = cc_series[cc_series["chart_id"] == chart_id]
            groups = {"primary": [], "secondary": []}
            for _, srow in rows.iterrows():
                key = (str(srow["haver_ticker"]).strip(), str(srow.get("transform", "level")).strip() or "level")
                if key not in colkey_index:
                    continue
                groups[str(srow.get("axis", "primary")).strip() or "primary"].append(
                    (colkey_index[key], str(srow.get("label", key[0])), str(srow.get("type", "line")))
                )
            if not groups["primary"] and not groups["secondary"]:
                continue
            if not groups["primary"]:
                groups["primary"], groups["secondary"] = groups["secondary"], []

            def make_chart(members):
                use_bar = all(t == "bar" for _, _, t in members) and members
                ch = BarChart() if use_bar else LineChart()
                if not use_bar:
                    ch.x_axis = DateAxis(crossAx=100)
                    ch.x_axis.number_format = "yyyy"
                    ch.x_axis.majorTimeUnit = "years"
                ch.y_axis.axId = 100
                for idx, label, _t in members:
                    col = idx + 2
                    ref = Reference(ws_d, min_col=col, min_row=first_data_row, max_row=last_data_row)
                    ser = Series(ref, title=None)
                    ser.tx = SeriesLabel(strRef=StrRef(f"'{cc}_data'!{get_column_letter(col)}1"))
                    ser.smooth = False
                    ch.series.append(ser)
                dates = Reference(ws_d, min_col=1, min_row=first_data_row, max_row=last_data_row)
                ch.set_categories(dates)
                return ch

            chart = make_chart(groups["primary"])
            if groups["secondary"]:
                chart2 = make_chart(groups["secondary"])
                chart2.y_axis.axId = 200
                chart2.y_axis.crosses = "max"
                chart += chart2

            title = str(chrow.get("title", "") or "").strip()
            if title:
                chart.title = title
            chart.height = 7.5
            chart.width = 15.0
            chart.legend.position = "b"

            anchor_row = (pos // 2) * CHART_H_ROWS + 2
            anchor_col = (pos % 2) * CHART_W_COLS + 2
            ws_c.add_chart(chart, f"{get_column_letter(anchor_col)}{anchor_row}")
            pos += 1
            n_charts_total += 1

        log.info("%s: 차트 %d개 / 시리즈 %d열", cc, pos, len(panel_cols))

    wb.save(OUTPUT_FILE)
    if skipped_series:
        for cc, hv, why in skipped_series:
            log.warning("시리즈 제외 %s %s (%s)", cc, hv, why)
    log.info("저장 완료: %s (차트 %d개)", OUTPUT_FILE.name, n_charts_total)
    return 0


if __name__ == "__main__":
    sys.exit(main())
